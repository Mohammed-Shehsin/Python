import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb =xl.load_workbook(filename) #fILE NAME
    sheet=wb['Sheet 1']
    # cell=sheet.cell(1,1)   #ton access cell A1
    # print(sheet.max_row)   #To know the number of rows
    # print(sheet.max_column)   #To know the number of column
    for row in range(2,sheet.max_row+1):
        cell=sheet.cell(row,3)
        corrected_price=cell.value * 0.9
        corrected_price_cell=sheet(row,4)
        corrected_price_cell.value=corrected_price

    value=Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)
    bar_chart=BarChart()
    bar_chart.add_data(value)
    sheet.add_chart(bar_chart,'e2')


    wb.save(filename)


process_workbook(" ") #filename